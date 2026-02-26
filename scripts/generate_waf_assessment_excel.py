from __future__ import annotations

import csv
import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]


def _read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def _read_support_cases(csv_path: Path) -> list[dict[str, str]]:
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader)


def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_table(ws, table_name: str, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def build_workbook(
    *,
    architecture_md_path: Path,
    support_cases_csv_path: Path,
    output_xlsx_path: Path,
    assessment_date: dt.date,
) -> Path:
    architecture_md = _read_text(architecture_md_path)
    support_cases = _read_support_cases(support_cases_csv_path)

    wb = Workbook()
    wb.properties.title = "Azure WAF Assessment"
    wb.properties.created = dt.datetime.now(dt.timezone.utc)

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")

    # Sheet: Overview
    ws = wb.active
    ws.title = "Overview"
    ws.freeze_panes = "A6"

    ws["A1"] = "Well-Architected Framework (WAF) Assessment"
    ws["A1"].font = Font(size=16, bold=True)

    ws["A3"] = "Assessment Date"
    ws["B3"] = assessment_date.isoformat()
    ws["A4"] = "Architecture Summary"
    ws["B4"] = "Hub-and-spoke AKS with Application Gateway → ILB → Traefik, Private Link, Azure Firewall, Bastion, and Azure Monitor"

    for addr in ("A3", "A4"):
        ws[addr].font = Font(bold=True)

    ws["A6"] = "Source Architecture Document (excerpt)"
    ws["A6"].font = Font(bold=True)
    ws["A7"] = architecture_md
    ws["A7"].alignment = Alignment(wrap_text=True, vertical="top")

    _set_col_widths(ws, {1: 28, 2: 120})
    ws.row_dimensions[7].height = 320

    # Sheet: Scores
    ws = wb.create_sheet("Scores")
    ws.freeze_panes = "A2"

    ws.append(["Pillar", "Score", "Status"])
    for c in range(1, 4):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    scores = [
        ("Reliability", 68, "Fair"),
        ("Security", 74, "Good"),
        ("Cost Optimization", 64, "Fair"),
        ("Operational Excellence", 66, "Fair"),
        ("Performance Efficiency", 78, "Good"),
        ("Overall", 70, "Good baseline, needs hardening"),
    ]
    for row in scores:
        ws.append(list(row))

    _set_col_widths(ws, {1: 28, 2: 10, 3: 35})
    for r in range(2, 2 + len(scores)):
        ws.cell(row=r, column=1).alignment = top
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3).alignment = top

    # Conditional formatting for score
    ws.conditional_formatting.add(
        f"B2:B{1 + len(scores)}",
        CellIsRule(operator="lessThan", formula=["60"], fill=PatternFill("solid", fgColor="F8D7DA")),
    )
    ws.conditional_formatting.add(
        f"B2:B{1 + len(scores)}",
        CellIsRule(operator="between", formula=["60", "79"], fill=PatternFill("solid", fgColor="FFF3CD")),
    )
    ws.conditional_formatting.add(
        f"B2:B{1 + len(scores)}",
        CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=PatternFill("solid", fgColor="D1E7DD")),
    )

    _apply_table(ws, "ScoresTable", 1, 1, 1 + len(scores), 3)

    # Sheet: Recommendations
    ws = wb.create_sheet("Recommendations")
    ws.freeze_panes = "A2"

    rec_headers = [
        "ID",
        "Recommendation",
        "Primary Pillar(s)",
        "Why / Evidence",
        "Risk if not addressed",
        "Priority",
        "Estimated Effort",
        "Related Support Cases",
        "Key Docs",
    ]
    ws.append(rec_headers)
    for c in range(1, len(rec_headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    recommendations = [
        {
            "id": "REC-01",
            "title": "Standardize ingress routing and health probes end-to-end (App Gateway → ILB → Traefik)",
            "pillars": "Reliability, Operational Excellence",
            "why": "Repeated ingress/probe/routing misconfig issues leading to 502s and backend unhealthy signals.",
            "risk": "Intermittent or complete service outage; slow triage due to multi-layer routing.",
            "priority": "High",
            "effort": "1–3 days",
            "cases": "120012346; 120012353",
            "docs": "https://learn.microsoft.com/azure/application-gateway/application-gateway-probe-overview | https://learn.microsoft.com/azure/architecture/reference-architectures/containers/aks-baseline",
        },
        {
            "id": "REC-02",
            "title": "Automate TLS certificate lifecycle for Application Gateway using Key Vault",
            "pillars": "Reliability, Security",
            "why": "Certificate expiry caused TLS handshake failures and outage of public endpoints.",
            "risk": "Public endpoint downtime and incident escalation during renewal windows.",
            "priority": "Critical",
            "effort": "0.5–2 days",
            "cases": "120012347",
            "docs": "https://learn.microsoft.com/azure/key-vault/certificates/about-certificates | https://learn.microsoft.com/azure/application-gateway/key-vault-certs",
        },
        {
            "id": "REC-03",
            "title": "Harden Private Link + Private DNS zone management (Key Vault, ACR)",
            "pillars": "Security, Reliability",
            "why": "Private Endpoint access failures caused secret retrieval and DNS resolution issues.",
            "risk": "Workloads unable to pull images or read secrets; cascading failures.",
            "priority": "High",
            "effort": "1–2 days",
            "cases": "120012348",
            "docs": "https://learn.microsoft.com/azure/private-link/private-endpoint-dns | https://learn.microsoft.com/azure/key-vault/general/private-link-service",
        },
        {
            "id": "REC-04",
            "title": "Make monitoring egress explicit and continuously validated through Azure Firewall",
            "pillars": "Operational Excellence, Reliability",
            "why": "Firewall misconfig blocked Azure Monitor/Prometheus endpoints; telemetry gaps.",
            "risk": "Reduced observability during incidents; delayed detection of failures.",
            "priority": "High",
            "effort": "1–3 days",
            "cases": "120012350; 120012354",
            "docs": "https://learn.microsoft.com/azure/firewall/overview | https://learn.microsoft.com/azure/azure-monitor/containers/kubernetes-monitoring-enable",
        },
        {
            "id": "REC-05",
            "title": "Implement AKS subnet/IP capacity planning and autoscaler guardrails",
            "pillars": "Reliability, Performance Efficiency",
            "why": "Autoscaling failures due to subnet IP range constraints.",
            "risk": "Scale-out fails; pods stuck Pending; degraded performance/availability.",
            "priority": "High",
            "effort": "0.5–2 days",
            "cases": "120012351",
            "docs": "https://learn.microsoft.com/azure/aks/cluster-autoscaler | https://learn.microsoft.com/azure/aks/concepts-network",
        },
        {
            "id": "REC-06",
            "title": "Reduce configuration drift via IaC + policy guardrails + runbooks",
            "pillars": "Operational Excellence, Security",
            "why": "Pattern of incidents rooted in configuration mistakes (Bastion association, probes, routing, rules).",
            "risk": "Recurring outages; inconsistent environments; slower recovery.",
            "priority": "High",
            "effort": "1–3 weeks (incremental)",
            "cases": "120012352; 120012346; 120012353; 120012350",
            "docs": "https://learn.microsoft.com/azure/governance/policy/overview | https://learn.microsoft.com/azure/well-architected/",
        },
    ]

    for rec in recommendations:
        ws.append(
            [
                rec["id"],
                rec["title"],
                rec["pillars"],
                rec["why"],
                rec["risk"],
                rec["priority"],
                rec["effort"],
                rec["cases"],
                rec["docs"],
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=1 + len(recommendations), min_col=1, max_col=len(rec_headers)):
        for cell in row:
            cell.alignment = wrap

    _set_col_widths(
        ws,
        {
            1: 10,
            2: 46,
            3: 26,
            4: 38,
            5: 34,
            6: 10,
            7: 16,
            8: 22,
            9: 60,
        },
    )
    _apply_table(ws, "RecommendationsTable", 1, 1, 1 + len(recommendations), len(rec_headers))

    # Sheet: Roadmap
    ws = wb.create_sheet("Roadmap")
    ws.freeze_panes = "A2"
    ws.append(["Horizon", "Focus", "Actions"])
    for c in range(1, 4):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    roadmap = [
        (
            "Next 7 days",
            "Stability quick wins",
            "Automate cert renewal + alerts (REC-02); standardize probes/routing (REC-01); validate monitoring egress (REC-04)",
        ),
        (
            "Next 30 days",
            "Platform hardening",
            "Automate Private DNS/Private Link zone linking (REC-03); add subnet/IP guardrails and alerts (REC-05); add telemetry-drop alerting (REC-04)",
        ),
        (
            "Next 90 days",
            "Operational maturity",
            "Expand IaC coverage and policy guardrails; publish runbooks and do game-day drills (REC-06)",
        ),
    ]
    for r in roadmap:
        ws.append(list(r))

    for row in ws.iter_rows(min_row=2, max_row=1 + len(roadmap), min_col=1, max_col=3):
        for cell in row:
            cell.alignment = wrap

    _set_col_widths(ws, {1: 14, 2: 22, 3: 110})
    _apply_table(ws, "RoadmapTable", 1, 1, 1 + len(roadmap), 3)

    # Sheet: Support Cases
    ws = wb.create_sheet("SupportCases")
    ws.freeze_panes = "A2"

    if support_cases:
        headers = list(support_cases[0].keys())
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        for item in support_cases:
            ws.append([item.get(h, "") for h in headers])

        for row in ws.iter_rows(min_row=2, max_row=1 + len(support_cases), min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = wrap

        widths = {
            1: 12,  # ticketnumber
            2: 44,  # title
            3: 12,  # createdon
            4: 28,  # product
            5: 60,  # root cause
            6: 56,  # customer statement
            7: 56,  # resolution
            8: 28,  # fullpath
        }
        _set_col_widths(ws, widths)
        _apply_table(ws, "SupportCasesTable", 1, 1, 1 + len(support_cases), len(headers))
        ws.auto_filter.ref = ws.dimensions

    output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx_path)
    return output_xlsx_path


def main() -> None:
    assessment_date = dt.date.today()
    architecture_md_path = ROOT / ".github" / "skills" / "waf-assessment" / "mid" / "architecture_document.md"
    support_cases_csv_path = ROOT / ".github" / "skills" / "waf-assessment" / "mid" / "azure_support_cases.csv"
    output_xlsx_path = ROOT / ".github" / "skills" / "waf-assessment" / "mid" / "waf_assessment_results.xlsx"

    if not architecture_md_path.exists():
        raise FileNotFoundError(f"Missing architecture doc: {architecture_md_path}")
    if not support_cases_csv_path.exists():
        raise FileNotFoundError(f"Missing support cases CSV: {support_cases_csv_path}")

    out = build_workbook(
        architecture_md_path=architecture_md_path,
        support_cases_csv_path=support_cases_csv_path,
        output_xlsx_path=output_xlsx_path,
        assessment_date=assessment_date,
    )
    print(str(out))


if __name__ == "__main__":
    main()
